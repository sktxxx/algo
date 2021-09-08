from selenium import webdriver
from openpyxl import load_workbook
import time

def init_driver():
    driver=webdriver.Chrome()
    driver.implicitly_wait(5)
    return driver

def read_excel():
    wb = load_workbook("mall.xlsx")
    ws = wb.active
    list_jar = []
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=2,values_only=True):
        list_jar.append(row)
    return list_jar


def input_form(jar):
    driver = webdriver.Chrome()
    driver.get('https://www.baidu.com')
    driver.find_element_by_class_name('s_ipt').send_keys(jar[0])
    driver.find_element_by_id('su').click()
    time.sleep(3)



if __name__ == '__main__':

    # 初始化谷歌浏览器驱动器
    driver = init_driver()

    # 读取Excel 文件，把每行数据放到列表里。
    list_jar = read_excel()

    # 遍历列表，取出用户信息
    for jar in list_jar:
        # 执行填写表单操作
        input_form(jar)
    # 最后退出浏览器
    driver.quit()



